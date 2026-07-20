import customtkinter as ctk
import json
from datetime import datetime
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from ui.theme import COLORS, THEME_PALETTES
from ui.components import AppCard, PageTitle, PageSubtitle, PrimaryButton, AppEntry
from core.alarm_sounds import ALARM_SOUNDS

class SettingsPage(ctk.CTkFrame):
    def __init__(self, parent, app):
        super().__init__(parent, fg_color=COLORS["bg"])
        self.app = app
        self.pending_reset_action = None
        self.pending_import_file_path = None

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.scroll = ctk.CTkScrollableFrame(
            self,
            fg_color=COLORS["bg"],
            corner_radius=0
        )
        self.scroll.grid(row=0, column=0, sticky="nsew")
        self.scroll.grid_columnconfigure(0, weight=1)

        self.create_header()
        self.create_settings_card()
        self.load_settings()

    def create_header(self):
        self.header = ctk.CTkFrame(self.scroll, fg_color="transparent")
        self.header.grid(row=0, column=0, padx=32, pady=(28, 18), sticky="ew")
        self.header.grid_columnconfigure(0, weight=1)

        self.title_label = ctk.CTkLabel(
            self.header,
            text=self.app.t("settings"),
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=30, weight="bold"),
            anchor="w"
        )
        self.title_label.grid(row=0, column=0, sticky="w")

        self.subtitle_label = ctk.CTkLabel(
            self.header,
            text=self.app.t("settings_subtitle"),
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=14),
            anchor="w"
        )
        self.subtitle_label.grid(row=1, column=0, pady=(6, 0), sticky="w")

    def preview_and_save_alarm(self, selected_name):
        # Menüde görünen çevrilmiş isimden sabit alarm ID'sini bul.
        alarm_id = self.alarm_display_names.get(selected_name)

        if not alarm_id:
            print(f"Alarm ID bulunamadı: {selected_name}")
            return

        # Sabit alarm ID'sini kaydet.
        settings = self.get_settings()
        settings["selected_alarm"] = alarm_id
        self.app.save_app_data()

        # Önizleme sesini merkezi alarm sistemi çalsın.
        self.app.preview_alarm(
            alarm_id=alarm_id,
            duration_ms=4000
        )

        self.status_label.configure(
            text=f"{selected_name} {self.app.t('selected')}",
            text_color=COLORS["green"]
        )

        self.after(
            2000,
            lambda: self.status_label.configure(text="")
        )

    def import_app_data(self):
        self.pending_reset_action = None

        if self.pending_import_file_path:
            self.confirm_import_app_data()
            return

        file_path = filedialog.askopenfilename(
            title=self.app.t("select_backup_file"),
            filetypes=[("JSON files", "*.json")]
        )

        if not file_path:
            return

        try:
            with open(file_path, "r", encoding="utf-8") as file:
                imported_data = json.load(file)

            self.validate_imported_data(imported_data)

            self.pending_import_file_path = file_path

            self.status_label.configure(
                text=self.app.t("confirm_import_data"),
                text_color=COLORS["orange"]
            )

            self.after(7000, self.clear_pending_import)

        except Exception:
            self.pending_import_file_path = None
            self.status_label.configure(
                text=self.app.t("data_import_failed"),
                text_color=COLORS["red"]
            )
            self.after(3000, lambda: self.status_label.configure(text=""))

    def validate_imported_data(self, imported_data):
        if not isinstance(imported_data, dict):
            raise ValueError("Invalid data format")

        required_keys = ["settings", "tasks", "subjects", "sessions"]

        for key in required_keys:
            if key not in imported_data:
                raise ValueError("Missing FocusFlow data keys")

        if not isinstance(imported_data.get("settings"), dict):
            raise ValueError("Invalid settings format")

        if not isinstance(imported_data.get("tasks"), list):
            raise ValueError("Invalid tasks format")

        if not isinstance(imported_data.get("subjects"), list):
            raise ValueError("Invalid subjects format")

        if not isinstance(imported_data.get("sessions"), list):
            raise ValueError("Invalid sessions format")
        
    def confirm_import_app_data(self):
        file_path = self.pending_import_file_path

        if not file_path:
            return

        try:
            with open(file_path, "r", encoding="utf-8") as file:
                imported_data = json.load(file)

            self.validate_imported_data(imported_data)

            self.app.app_data.clear()
            self.app.app_data.update(imported_data)

            if hasattr(self.app, "ensure_app_data_defaults"):
                self.app.ensure_app_data_defaults()
            else:
                self.app.save_app_data()

            self.pending_import_file_path = None
            self.refresh_after_import()

            self.status_label.configure(
                text=self.app.t("data_imported"),
                text_color=COLORS["green"]
            )

        except Exception:
            self.pending_import_file_path = None
            self.status_label.configure(
                text=self.app.t("data_import_failed"),
                text_color=COLORS["red"]
            )

        self.after(3000, lambda: self.status_label.configure(text=""))

    def clear_pending_import(self):
        self.pending_import_file_path = None

        if self.pending_reset_action is None:
            self.status_label.configure(text="")
    
    def refresh_after_import(self):
        self.load_settings()

        if hasattr(self.app, "todo_page"):
            if hasattr(self.app.todo_page, "refresh_subject_menu"):
                self.app.todo_page.refresh_subject_menu()
            self.app.todo_page.render_tasks()

        if hasattr(self.app, "focus_page"):
            self.app.focus_page.load_active_task()
            self.app.focus_page.update_queue_progress()
            self.app.focus_page.refresh_queue_progress_visibility()
            if hasattr(self.app.focus_page, "refresh_away_card_visibility"):
                self.app.focus_page.refresh_away_card_visibility()

        if hasattr(self.app, "pomodoro_page"):
            self.app.pomodoro_page.load_pomodoro_settings()
            self.app.pomodoro_page.populate_settings_entries()
            self.app.pomodoro_page.update_timer_label()
            self.app.pomodoro_page.update_mode_ui()
            self.app.pomodoro_page.update_cycle_labels()
            self.app.pomodoro_page.update_session_info_values()
            self.app.pomodoro_page.update_auto_start_info()

        if hasattr(self.app, "subjects_page"):
            self.app.subjects_page.render_subjects()

        if hasattr(self.app, "statistics_page"):
            if hasattr(self.app.statistics_page, "refresh_subject_filter_menu"):
                self.app.statistics_page.refresh_subject_filter_menu()
            self.app.statistics_page.refresh_stats()
    
    def create_settings_card(self):
        self.settings_card = AppCard(self.scroll)
        self.settings_card.grid(row=1, column=0, padx=36, pady=(12, 30), sticky="ew")
        self.settings_card.grid_columnconfigure(0, weight=1)

        self.week_start_options = {
            "monday": self.app.t("week_start_monday"),
            "sunday": self.app.t("week_start_sunday")
}

        self.auto_focus_frame = self.create_setting_row(
            row=0,
            title_key="auto_start_focus",
            description_key="auto_start_focus_desc"
        )

        self.auto_focus_switch = ctk.CTkSwitch(
            self.auto_focus_frame,
            text="",
            progress_color=COLORS["primary"],
            button_color=COLORS["text"],
            button_hover_color=COLORS["soft"],
            command=self.save_settings
        )
        self.auto_focus_switch.grid(row=0, column=1, rowspan=2, padx=20, pady=18)

        self.auto_break_frame = self.create_setting_row(
            row=1,
            title_key="auto_start_break",
            description_key="auto_start_break_desc"
        )

        self.auto_break_switch = ctk.CTkSwitch(
            self.auto_break_frame,
            text="",
            progress_color=COLORS["primary"],
            button_color=COLORS["text"],
            button_hover_color=COLORS["soft"],
            command=self.save_settings
        )
        self.auto_break_switch.grid(row=0, column=1, rowspan=2, padx=20, pady=18)
        
        self.sound_frame = self.create_setting_row(
            row=2,
            title_key="sound_notification",
            description_key="sound_notification_desc"
        )

        self.sound_switch = ctk.CTkSwitch(
            self.sound_frame,
            text="",
            progress_color=COLORS["primary"],
            button_color=COLORS["text"],
            button_hover_color=COLORS["soft"],
            command=self.save_settings
        )
        self.sound_switch.grid(row=0, column=1, rowspan=2, padx=20, pady=18)

        self.alarm_display_names = {
            self.app.t(data["name_key"]): alarm_id
            for alarm_id, data in ALARM_SOUNDS.items()
        }

        # Alarm Seçim Frame'i
        self.alarm_frame = ctk.CTkFrame(self.settings_card, fg_color=COLORS["surface"], corner_radius=18)
        self.alarm_frame.grid(row=3, column=0, padx=20, pady=(8, 8), sticky="ew")
        self.alarm_frame.grid_columnconfigure(0, weight=1)

        self.alarm_label = ctk.CTkLabel(self.alarm_frame, text=self.app.t("alarm_sound"), text_color=COLORS["text"], font=ctk.CTkFont(size=15, weight="bold"))
        self.alarm_label.grid(row=0, column=0, padx=18, pady=(16, 2), sticky="w")

        self.alarm_desc = ctk.CTkLabel(self.alarm_frame, text=self.app.t("alarm_sound_desc"), text_color=COLORS["muted"], font=ctk.CTkFont(size=13))
        self.alarm_desc.grid(row=1, column=0, padx=18, pady=(0, 16), sticky="w")

        self.alarm_menu = ctk.CTkOptionMenu(
            self.alarm_frame, values=list(self.alarm_display_names.keys()), width=160, height=40,
            fg_color=COLORS["primary"], button_color=COLORS["primary"],
            button_hover_color=COLORS["primary_hover"], text_color=COLORS["white"],
            dropdown_fg_color=COLORS["surface"], dropdown_text_color=COLORS["text"], 
            command=self.preview_and_save_alarm
        )
        self.alarm_menu.grid(row=0, column=1, rowspan=2, padx=20, pady=18, sticky="e")

        self.goal_frame = ctk.CTkFrame(
            self.settings_card,
            fg_color=COLORS["surface"],
            corner_radius=18
        )
        self.goal_frame.grid(
            row=4,
            column=0,
            padx=20,
            pady=(8, 8),
            sticky="ew"
        )
        self.goal_frame.grid_columnconfigure(0, weight=1)
        self.goal_frame.grid_columnconfigure(1, weight=0)
        self.goal_frame.grid_columnconfigure(2, weight=0)

        self.goal_title = ctk.CTkLabel(
            self.goal_frame,
            text=self.app.t("daily_focus_goal"),
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=15, weight="bold")
        )
        self.goal_title.grid(row=0, column=0, padx=18, pady=(16, 2), sticky="w")

        self.goal_desc = ctk.CTkLabel(
            self.goal_frame,
            text=self.app.t("daily_focus_goal_desc"),
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=13)
        )
        self.goal_desc.grid(row=1, column=0, padx=18, pady=(0, 16), sticky="w")

        self.goal_entry = AppEntry(
            self.goal_frame,
            placeholder_text=self.app.t("minutes_short"),
            width=90
        )
        self.goal_entry.grid(row=0, column=1, rowspan=2, padx=(12, 8), pady=18, sticky="e")
        self.save_button = PrimaryButton(
            self.goal_frame,
            text=self.app.t("save"),
            command=self.save_settings,
            width=90
        )
        self.save_button.grid(row=0, column=2, rowspan=2, padx=(0, 18), pady=18, sticky="e")

        self.queue_progress_frame = self.create_setting_row(
            row=5,
            title_key="show_queue_progress",
            description_key="show_queue_progress_desc"
        )

        self.queue_progress_switch = ctk.CTkSwitch(
            self.queue_progress_frame,
            text="",
            progress_color=COLORS["primary"],
            button_color=COLORS["text"],
            button_hover_color=COLORS["soft"],
            command=self.save_settings
        )
        self.queue_progress_switch.grid(row=0, column=1, rowspan=2, padx=20, pady=18)

        self.cumulative_away_frame = self.create_setting_row(
            row=6,
            title_key="show_cumulative_away_time",
            description_key="show_cumulative_away_time_desc"
        )

        self.cumulative_away_switch = ctk.CTkSwitch(
            self.cumulative_away_frame,
            text="",
            progress_color=COLORS["primary"],
            button_color=COLORS["text"],
            button_hover_color=COLORS["soft"],
            command=self.save_settings
        )
        self.cumulative_away_switch.grid(row=0, column=1, rowspan=2, padx=20, pady=18)

        self.week_start_frame = ctk.CTkFrame(
            self.settings_card,
            fg_color=COLORS["surface"],
            corner_radius=18
        )
        self.week_start_frame.grid(row=7, column=0, padx=20, pady=(8, 8), sticky="ew")
        self.week_start_frame.grid_columnconfigure(0, weight=1)

        self.week_start_label = ctk.CTkLabel(
            self.week_start_frame,
            text=self.app.t("week_start_day"),
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=15, weight="bold")
        )
        self.week_start_label.grid(row=0, column=0, padx=18, pady=(16, 2), sticky="w")

        self.week_start_desc = ctk.CTkLabel(
            self.week_start_frame,
            text=self.app.t("week_start_day_desc"),
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=13),
            wraplength=520,
            justify="left"
        )
        self.week_start_desc.grid(row=1, column=0, padx=18, pady=(0, 16), sticky="w")

        self.week_start_menu = ctk.CTkOptionMenu(
            self.week_start_frame,
            values=[
                self.app.t("week_start_monday"),
                self.app.t("week_start_sunday")
            ],
            width=160, height=40,
            fg_color=COLORS["primary"], button_color=COLORS["primary"],
            button_hover_color=COLORS["primary_hover"], text_color=COLORS["white"],
            dropdown_fg_color=COLORS["surface"], dropdown_text_color=COLORS["text"],
            command=self.change_week_start_day
        )
        self.week_start_menu.grid(row=0, column=1, rowspan=2, padx=20, pady=18, sticky="e")
        
        self.palette_frame = ctk.CTkFrame(
            self.settings_card,
            fg_color=COLORS["surface"],
            corner_radius=18
        )
        self.palette_frame.grid(row=8, column=0, padx=20, pady=(8, 8), sticky="ew")
        self.palette_frame.grid_columnconfigure(0, weight=1)

        self.palette_title = ctk.CTkLabel(
            self.palette_frame, text=self.app.t("color_palette"), text_color=COLORS["text"],
            font=ctk.CTkFont(size=15, weight="bold")
        )
        self.palette_title.grid(row=0, column=0, padx=18, pady=(16, 2), sticky="w")
        self.palette_desc = ctk.CTkLabel(
            self.palette_frame, text=self.app.t("color_palette_desc"), text_color=COLORS["muted"],
            font=ctk.CTkFont(size=13), wraplength=520, justify="left"
        )
        self.palette_desc.grid(row=1, column=0, padx=18, pady=(0, 16), sticky="w")
        self.palette_values = {data["name"]: key for key, data in THEME_PALETTES.items()}
        self.palette_menu = ctk.CTkOptionMenu(
            self.palette_frame, values=list(self.palette_values), width=160, height=40,
            fg_color=COLORS["primary"], button_color=COLORS["primary"],
            button_hover_color=COLORS["primary_hover"], text_color=COLORS["white"],
            dropdown_fg_color=COLORS["surface"], dropdown_text_color=COLORS["text"],
            command=self.change_color_palette
        )
        self.palette_menu.grid(row=0, column=1, rowspan=2, padx=20, pady=18, sticky="e")

        self.data_frame = ctk.CTkFrame(
            self.settings_card,
            fg_color=COLORS["surface"],
            corner_radius=18
        )
        self.data_frame.grid(row=9, column=0, padx=20, pady=(8, 20), sticky="ew")
        self.data_frame.grid_columnconfigure(0, weight=1)
        self.data_frame.grid_columnconfigure(1, weight=1)

        self.data_title = ctk.CTkLabel(
            self.data_frame,
            text=self.app.t("data_management"),
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=15, weight="bold")
        )
        self.data_title.grid(row=0, column=0, columnspan=2, padx=18, pady=(16, 2), sticky="w")

        self.data_desc = ctk.CTkLabel(
            self.data_frame,
            text=self.app.t("data_management_desc"),
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=13),
            wraplength=520,
            justify="left"
        )
        self.data_desc.grid(row=1, column=0, columnspan=2, padx=18, pady=(0, 14), sticky="w")

        self.export_history_button = PrimaryButton(
            self.data_frame,
            text=self.app.t("export_study_history"),
            command=self.export_study_history_excel,
            width=140
        )
        self.export_history_button.grid(
            row=2,
            column=0,
            columnspan=2,
            padx=18,
            pady=(0, 10),
            sticky="ew"
        )
        self.reset_stats_button = PrimaryButton(
            self.data_frame,
            text=self.app.t("reset_statistics"),
            command=self.confirm_reset_statistics,
            width=140
        )
        self.reset_stats_button.grid(row=3, column=0, padx=(18, 8), pady=(0, 18), sticky="ew")

        self.reset_app_button = PrimaryButton(
            self.data_frame,
            text=self.app.t("reset_application"),
            command=self.confirm_reset_application,
            width=140
        )
        self.reset_app_button.grid(row=3, column=1, padx=(8, 18), pady=(0, 18), sticky="ew")

        self.export_data_button = PrimaryButton(
            self.data_frame,
            text=self.app.t("export_data"),
            command=self.export_app_data,
            width=140
        )
        self.export_data_button.grid(row=4, column=0, padx=(18, 8), pady=(0, 10), sticky="ew")

        self.import_data_button = PrimaryButton(
            self.data_frame,
            text=self.app.t("import_data"),
            command=self.import_app_data,
            width=140
        )
        self.import_data_button.grid(row=4, column=1, padx=(8, 18), pady=(0, 10), sticky="ew")

        self.status_label = ctk.CTkLabel(
            self.settings_card,
            text="",
            text_color=COLORS["green"],
            font=ctk.CTkFont(size=13, weight="bold"))
        self.status_label.grid(row=10, column=0, padx=20, pady=(0, 18), sticky="w")

    def export_study_history_excel(self):
        sessions = self.app.app_data.get("sessions", [])

        if not sessions:
            self.status_label.configure(
                text=self.app.t("no_study_history"),
                text_color=COLORS["orange"]
            )
            self.after(3000, lambda: self.status_label.configure(text=""))
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"focusflow_study_history_{timestamp}.xlsx"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not file_path:
            return

        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Study History"

            headers = [
                "Date",
                "Time",
                "Task",
                "Subject",
                "Source",
                "Mode",
                "Focus Minutes",
                "Away Minutes",
                "Completed At",
                "Session ID"
            ]
            worksheet.append(headers)

            header_fill = PatternFill(fill_type="solid", fgColor="6D5DF6")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            sorted_sessions = sorted(
                sessions,
                key=lambda session: session.get("completed_at", "")
            )

            for session in sorted_sessions:
                completed_at = session.get("completed_at", "")
                date_text = ""
                time_text = ""

                if completed_at:
                    try:
                        parsed_datetime = datetime.fromisoformat(completed_at)
                        date_text = parsed_datetime.strftime("%Y-%m-%d")
                        time_text = parsed_datetime.strftime("%H:%M:%S")
                    except (TypeError, ValueError):
                        date_text = str(completed_at)

                duration_seconds = session.get("duration_seconds", 0) or 0
                away_seconds = session.get("away_seconds", 0) or 0

                task_title = session.get("task_title")
                if not task_title:
                    if session.get("source") == "regular_pomodoro":
                        task_title = self.app.t("regular_pomodoro_session")
                    else:
                        task_title = self.app.t("untitled_task")

                worksheet.append([
                    date_text,
                    time_text,
                    task_title,
                    session.get("subject_name", ""),
                    session.get("source", ""),
                    session.get("mode", ""),
                    round(duration_seconds / 60, 2),
                    round(away_seconds / 60, 2),
                    completed_at,
                    session.get("id", "")
                ])

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            column_widths = {1: 13, 2: 11, 3: 32, 4: 24, 5: 20, 6: 14, 7: 16, 8: 14, 9: 22, 10: 22}
            for column_index, width in column_widths.items():
                worksheet.column_dimensions[get_column_letter(column_index)].width = width

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            workbook.save(file_path)
            self.status_label.configure(
                text=self.app.t("study_history_exported"),
                text_color=COLORS["green"]
            )

        except Exception as error:
            print(f"[EXCEL EXPORT ERROR] {error}")
            self.status_label.configure(
                text=self.app.t("study_history_export_failed"),
                text_color=COLORS["red"]
            )

        self.after(3000, lambda: self.status_label.configure(text=""))

    def export_app_data(self):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"focusflow_backup_{timestamp}.json"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            initialfile=default_filename,
            filetypes=[("JSON files", "*.json")]
        )

        if not file_path:
            return

        try:
            with open(file_path, "w", encoding="utf-8") as file:
                json.dump(self.app.app_data, file, ensure_ascii=False, indent=4)

            self.status_label.configure(
                text=self.app.t("data_exported"),
                text_color=COLORS["green"]
            )

        except Exception:
            self.status_label.configure(
                text=self.app.t("data_export_failed"),
                text_color=COLORS["red"]
            )

        self.after(2500, lambda: self.status_label.configure(text=""))
    
    def reset_application(self):
        current_language = self.app.app_data.get("language", "tr")

        self.app.app_data.clear()

        self.app.app_data.update({
            "language": current_language,
            "active_task_id": None,
            "queue_mode_active": False,
            "queue_task_ids": [],
            "last_queue_state": None,
            "settings": {
                "auto_start_break": False,
                "auto_start_focus": False,
                "sound_enabled": True,
                "daily_focus_goal_minutes": 300,
                "regular_focus_minutes": 25,
                "regular_short_break_minutes": 5,
                "regular_long_break_minutes": 15,
                "regular_long_break_after": 4,
                "regular_focus_count": 4,
                "show_queue_progress": True,
                "show_cumulative_away_time": True,
                "week_start_day": "monday",
                "appearance_mode": "dark",
                "color_palette": "purple"
            },
            "subjects": [
                {
                    "id": "subject_other",
                    "name": self.app.t("other_subject"),
                    "color": "#A78BFA",
                    "is_default": True
                }
            ],
            "tasks": [],
            "sessions": []
        })

        self.app.save_app_data()
        self.pending_reset_action = None

        self.load_settings()

        if hasattr(self.app, "todo_page"):
            if hasattr(self.app.todo_page, "refresh_subject_menu"):
                self.app.todo_page.refresh_subject_menu()
            self.app.todo_page.render_tasks()

        if hasattr(self.app, "focus_page"):
            self.app.focus_page.load_active_task()
            self.app.focus_page.update_queue_progress()
            self.app.focus_page.refresh_queue_progress_visibility()

        if hasattr(self.app, "pomodoro_page"):
            self.app.pomodoro_page.load_pomodoro_settings()
            self.app.pomodoro_page.populate_settings_entries()
            self.app.pomodoro_page.update_timer_label()
            self.app.pomodoro_page.update_mode_ui()
            self.app.pomodoro_page.update_cycle_labels()
            self.app.pomodoro_page.update_session_info_values()
            self.app.pomodoro_page.update_auto_start_info()

        if hasattr(self.app, "subjects_page"):
            self.app.subjects_page.render_subjects()

        if hasattr(self.app, "statistics_page"):
            self.app.statistics_page.refresh_subject_filter_menu()
            self.app.statistics_page.refresh_stats()

        self.status_label.configure(
            text=self.app.t("application_reset_done"),
            text_color=COLORS["green"]
        )

        self.after(2500, lambda: self.status_label.configure(text=""))

    def confirm_reset_application(self):
        if self.pending_reset_action != "application":
            self.pending_reset_action = "application"

            self.status_label.configure(
                text=self.app.t("confirm_reset_application"),
                text_color=COLORS["red"]
            )

            self.after(5000, self.clear_pending_reset_action)
            return

        self.reset_application()

    def confirm_reset_statistics(self):
        if self.pending_reset_action != "statistics":
            self.pending_reset_action = "statistics"

            self.status_label.configure(
                text=self.app.t("confirm_reset_statistics"),
                text_color=COLORS["orange"]
            )

            self.after(4000, self.clear_pending_reset_action)
            return

        self.reset_statistics()
    
    def clear_pending_reset_action(self):
        self.pending_reset_action = None
        self.status_label.configure(text="")

    def reset_statistics(self):
        self.app.app_data["sessions"] = []
        self.app.save_app_data()

        self.pending_reset_action = None

        if hasattr(self.app, "statistics_page"):
            self.app.statistics_page.refresh_stats()

        self.status_label.configure(
            text=self.app.t("statistics_reset_done"),
            text_color=COLORS["green"]
        )

        self.after(2500, lambda: self.status_label.configure(text=""))

    def create_setting_row(self, row, title_key, description_key):
        frame = ctk.CTkFrame(
            self.settings_card,
            fg_color=COLORS["surface"],
            corner_radius=18
        )
        frame.grid(row=row, column=0, padx=20, pady=(18 if row == 0 else 8, 8), sticky="ew")
        frame.grid_columnconfigure(0, weight=1)

        title = ctk.CTkLabel(
            frame,
            text=self.app.t(title_key),
            text_color=COLORS["text"],
            font=ctk.CTkFont(size=15, weight="bold")
        )
        title.grid(row=0, column=0, padx=18, pady=(16, 2), sticky="w")

        desc = ctk.CTkLabel(
            frame,
            text=self.app.t(description_key),
            text_color=COLORS["muted"],
            font=ctk.CTkFont(size=13),
            wraplength=520,
            justify="left"
        )
        desc.grid(row=1, column=0, padx=18, pady=(0, 16), sticky="w")

        frame.title_label = title
        frame.desc_label = desc
        frame.title_key = title_key
        frame.description_key = description_key

        return frame

    def get_settings(self):
        return self.app.app_data.setdefault("settings", {})

    def load_settings(self):
        settings = self.get_settings()

        if settings.get("auto_start_break", False):
            self.auto_break_switch.select()
        else:
            self.auto_break_switch.deselect()

        if settings.get("auto_start_focus", False):
            self.auto_focus_switch.select()
        else:
            self.auto_focus_switch.deselect()

        if settings.get("sound_enabled", True):
            self.sound_switch.select()
        else:
            self.sound_switch.deselect()

        selected_alarm = settings.get("selected_alarm", "Birdy")
        self.alarm_menu.set(selected_alarm)
        
        if settings.get("show_queue_progress", True):
            self.queue_progress_switch.select()
        else:
            self.queue_progress_switch.deselect()

        if settings.get("show_cumulative_away_time", True):
            self.cumulative_away_switch.select()
        else:
            self.cumulative_away_switch.deselect()

        palette_key = settings.get("color_palette", "purple")
        palette_name = THEME_PALETTES.get(palette_key, THEME_PALETTES["purple"])["name"]
        self.palette_menu.set(palette_name)

        settings = self.app.app_data.setdefault("settings", {})
        week_start_day = settings.get("week_start_day", "monday")

        if week_start_day == "sunday":
            self.week_start_menu.set(self.app.t("week_start_sunday"))
        else:
            self.week_start_menu.set(self.app.t("week_start_monday"))

        goal = settings.get("daily_focus_goal_minutes", 300)
        self.goal_entry.delete(0, "end")
        self.goal_entry.insert(0, str(goal))

    def change_color_palette(self, selected_name):
        palette_key = self.palette_values.get(selected_name, "purple")
        self.app.apply_theme(palette_key=palette_key)

    def change_week_start_day(self, selected_value):
        settings = self.app.app_data.setdefault("settings", {})

        if selected_value == self.app.t("week_start_sunday"):
            settings["week_start_day"] = "sunday"
        else:
            settings["week_start_day"] = "monday"

        self.app.save_app_data()

        if hasattr(self.app, "statistics_page"):
            self.app.statistics_page.refresh_stats()

    def save_settings(self):
        settings = self.get_settings()

        settings["auto_start_break"] = bool(self.auto_break_switch.get())
        settings["auto_start_focus"] = bool(self.auto_focus_switch.get())
        settings["sound_enabled"] = bool(self.sound_switch.get())

        goal_value = self.goal_entry.get().strip()

        try:
            goal_minutes = int(goal_value)
            if goal_minutes > 0:
                settings["daily_focus_goal_minutes"] = goal_minutes
        except ValueError:
            pass

        settings["show_queue_progress"] = bool(self.queue_progress_switch.get())
        settings["show_cumulative_away_time"] = bool(self.cumulative_away_switch.get())

        if hasattr(self.app, "focus_page"):
            self.app.focus_page.refresh_queue_progress_visibility()
            self.app.focus_page.refresh_away_card_visibility()

        self.app.save_app_data()
        self.status_label.configure(text=self.app.t("settings_saved"))

        self.after(2000, lambda: self.status_label.configure(text=""))

    def refresh_texts(self):
        self.title_label.configure(text=self.app.t("settings"))
        self.subtitle_label.configure(text=self.app.t("settings_subtitle"))

        for frame in [
            self.auto_break_frame,
            self.auto_focus_frame,
            self.sound_frame,
            self.queue_progress_frame,
            self.cumulative_away_frame
        ]:
            frame.title_label.configure(text=self.app.t(frame.title_key))
            frame.desc_label.configure(text=self.app.t(frame.description_key))

        self.alarm_label.configure(text=self.app.t("alarm_sound"))
        self.alarm_desc.configure(text=self.app.t("alarm_sound_desc"))

        self.week_start_label.configure(text=self.app.t("week_start_day"))
        self.week_start_desc.configure(text=self.app.t("week_start_day_desc"))
        self.palette_title.configure(text=self.app.t("color_palette"))
        self.palette_desc.configure(text=self.app.t("color_palette_desc"))

        settings = self.app.app_data.setdefault("settings", {})
        week_start_day = settings.get("week_start_day", "monday")

        self.week_start_menu.configure(
            values=[
                self.app.t("week_start_monday"),
                self.app.t("week_start_sunday")
            ]
        )

        if week_start_day == "sunday":
            self.week_start_menu.set(self.app.t("week_start_sunday"))
        else:
            self.week_start_menu.set(self.app.t("week_start_monday"))

        self.goal_title.configure(text=self.app.t("daily_focus_goal"))
        self.goal_desc.configure(text=self.app.t("daily_focus_goal_desc"))
        self.goal_entry.configure(placeholder_text=self.app.t("minutes_short"))
        self.save_button.configure(text=self.app.t("save"))
        self.data_title.configure(text=self.app.t("data_management"))
        self.data_desc.configure(text=self.app.t("data_management_desc"))
        self.export_data_button.configure(text=self.app.t("export_data"))
        self.import_data_button.configure(text=self.app.t("import_data"))
        self.export_history_button.configure(text=self.app.t("export_study_history"))
        self.reset_stats_button.configure(text=self.app.t("reset_statistics"))
        self.reset_app_button.configure(text=self.app.t("reset_application"))
